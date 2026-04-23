from flask import Flask, render_template, request, jsonify, send_from_directory
import pandas as pd
import datetime
import os
import qrcode
import calendar
import smtplib
from email.message import EmailMessage

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

app = Flask(__name__)

# ===== LOGIN =====
TEACHER_USERNAME = "Admin"
TEACHER_PASSWORD = "admin@123"

# ===== EMAIL =====
SENDER_EMAIL = "gamercode669@gmail.com"
APP_PASSWORD = "fhioshanbbbihegf"

# ===== LOAD STUDENTS =====
data = pd.read_excel("students.xlsx")
attendance = {}

# ===== COLORS =====
green = PatternFill(start_color="90EE90", fill_type="solid")
blue = PatternFill(start_color="87CEFA", fill_type="solid")
red = PatternFill(start_color="FF9999", fill_type="solid")

# ===== ROUTES =====
@app.route("/")
def home():
    return render_template("index.html")

@app.route("/scanner")
def scanner():
    return render_template("scanner.html")

@app.route("/success")
def success():
    return render_template("success.html")

@app.route("/dashboard")
def login():
    return render_template("login.html")

@app.route("/dashboard-home")
def dashboard():
    return render_template("dashboard.html")

@app.route('/images/<path:filename>')
def images(filename):
    return send_from_directory('images', filename)

# ===== QR EMAIL =====
@app.route("/generate-qr", methods=["POST"])
def generate_qr():
    roll = int(request.json["roll"])
    student = data[data["Roll No."] == roll]

    if student.empty:
        return jsonify({"status": "error"})

    email = student.iloc[0]["Email"]

    file = f"qr_{roll}.png"
    qrcode.make(f"Roll:{roll}").save(file)

    msg = EmailMessage()
    msg["Subject"] = "Attendance QR"
    msg["From"] = SENDER_EMAIL
    msg["To"] = email
    msg.set_content("Scan this QR")

    try:
        with open(file, "rb") as f:
            msg.add_attachment(f.read(), maintype="image", subtype="png", filename=file)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(SENDER_EMAIL, APP_PASSWORD)
            smtp.send_message(msg)
    except:
        pass

    os.remove(file)
    return jsonify({"status": "sent"})

# ===== EXCEL SYSTEM (FIXED + CLEAN) =====
def update_excel(att):
    now = datetime.datetime.now()
    month = now.strftime("%B")
    year = now.year
    file = f"{month}_Attendance.xlsx"
    today = now.day

    # CREATE FILE
    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active

        ws.cell(1,1).value = "Name"
        ws.cell(1,1).font = Font(bold=True)

        days = calendar.monthrange(year, now.month)[1]

        for d in range(1, days+1):
            col = d+1
            cell = ws.cell(1,col)
            cell.value = f"{d:02d}-{month[:3]}"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

            # Sunday red
            if calendar.weekday(year, now.month, d) == 6:
                cell.fill = red

        # Names
        for i,row in data.iterrows():
            ws.cell(i+2,1).value = row["Name"]

        ws.freeze_panes = "B2"
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active

    # UPDATE TODAY
    for i in range(2, ws.max_row+1):
        name = ws.cell(i,1).value
        status = "A"

        for r,s in att.items():
            st = data[data["Roll No."]==r]
            if not st.empty and st.iloc[0]["Name"]==name:
                status = s

        cell = ws.cell(i, today+1)

        if status == "P":
            cell.value = "P"
            cell.fill = green

        elif status == "Late":
            cell.value = "Late"
            cell.fill = blue

        else:
            cell.value = "A"
            cell.fill = red

        cell.alignment = Alignment(horizontal="center")

    wb.save(file)

# ===== MARK ATTENDANCE =====
@app.route("/mark-attendance", methods=["POST"])
def mark():
    roll = int(request.json["roll"])
    qr = request.json["qr"]
    teacher = request.json.get("teacher", False)

    if f"Roll:{roll}" not in qr:
        return jsonify({"status":"error","msg":"Invalid QR"})

    if roll in attendance:
        return jsonify({"status":"error","msg":"Already marked"})

    student = data[data["Roll No."]==roll]
    name = student.iloc[0]["Name"]

    now = datetime.datetime.now().time()

    if now < datetime.time(9,5):
        status="P"
    elif now <= datetime.time(9,15):
        status="Late"
    else:
        if teacher:
            status="Late"
        else:
            return jsonify({"status":"permission"})

    attendance[roll]=status
    update_excel(attendance)

    return jsonify({
        "status":"success",
        "name":name,
        "roll":roll,
        "time":datetime.datetime.now().strftime("%H:%M:%S"),
        "att":status
    })

# ===== LOGIN =====
@app.route("/check-login", methods=["POST"])
def check():
    username = request.json.get("username")
    password = request.json.get("password")

    if username == TEACHER_USERNAME and password == TEACHER_PASSWORD:
        return jsonify({"status":"ok"})
    return jsonify({"status":"fail"})

# ===== DASHBOARD =====
@app.route("/dashboard-data")
def dash():
    file = datetime.datetime.now().strftime("%B")+"_Attendance.xlsx"

    if not os.path.exists(file):
        return jsonify({"today":{}, "monthly":[]})

    df = pd.read_excel(file)

    today = datetime.datetime.now().day
    col = f"{today:02d}-{datetime.datetime.now().strftime('%b')}"

    # TODAY COUNT
    present = (df[col] == "P").sum()
    late = (df[col] == "Late").sum()
    absent = (df[col] == "A").sum()

    # MONTHLY %
    monthly = []
    total_days = len(df.columns) - 1

    for i,row in df.iterrows():
        vals = list(row[1:])
        p = vals.count("P")
        l = vals.count("Late")

        percent = round(((p+l)/total_days)*100,2)

        monthly.append({
            "Name": row["Name"],
            "Percentage": percent
        })

    return jsonify({
        "today": {
            "present": int(present),
            "late": int(late),
            "absent": int(absent)
        },
        "monthly": monthly
    })

# ===== RUN =====
app.run(debug=True)